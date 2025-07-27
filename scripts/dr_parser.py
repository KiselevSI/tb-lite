#!/usr/bin/env python3
"""
Сбор признаков лекарственной устойчивости из TB‑Profiler JSON‑отчётов
с расширенными колонками для каждой группы лекарств в Excel со многоуровневой шапкой:
  <Drug>  →  gene_name | Mutation | Freq

Изменение по просьбе пользователя:
- Вместо жёлтой заливки для вариантов с confidence == "Uncertain significance"
  рядом со значением мутации в колонке **Mutation** дописывается метка
  "[Uncertain significance]".

Правила:
- Включаем варианты с confidence == "Assoc w R" ИЛИ "Uncertain significance".
- Для "Uncertain significance" в колонке Mutation после конкретной мутации
  добавляется строка " [Uncertain significance]".
- Частота берётся из dr_variants[i].freq.
- В Excel шапка из 2 строк, индекс НЕ записываем, пустой первый столбец отсутствует.

Выход: Excel (.xlsx).

Зависимости: pandas, openpyxl
  pip install pandas openpyxl

Примеры
-------
# вывод в Excel
python dr_parser_excel.py -o result.xlsx sample1.json sample2.json

# если -о не указан, сохранит result.xlsx в текущей директории
python dr_parser_excel.py sample.json
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Dict, List, Tuple, Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

DRUGS: List[str] = [
    "Rifampicin",
    "Isoniazid",
    "Ethambutol",
    "Pyrazinamide",
    "Moxifloxacin",
    "Levofloxacin",
    "Bedaquiline",
    "Delamanid",
    "Pretomanid",
    "Linezolid",
    "Streptomycin",
    "Amikacin",
    "Kanamycin",
    "Capreomycin",
    "Clofazimine",
    "Ethionamide",
    "Para-aminosalicylic_acid",
    "Cycloserine",
]

# соответствие «имя из JSON → имя колонки» (нижний регистр)
JSON2COL: Dict[str, str] = {d.lower(): d for d in DRUGS}

# итоговая структура колонок: двухуровневые заголовки для Excel
# Верхний уровень: название препарата (для колонки Sample — пусто)
# Нижний уровень: gene_name / Mutation / Freq
TOP: List[str] = ["", ""]
BOTTOM: List[str] = ["Sample", "Other Variants"]
for drug in DRUGS:
    TOP.extend([drug, drug, drug])
    BOTTOM.extend(["gene_name", "Mutation", "Freq"])

COLUMNS = pd.MultiIndex.from_arrays([TOP, BOTTOM])


def _to_str(x: Any) -> str:
    if x is None:
        return ""
    if isinstance(x, float):
        # компактное представление дробей
        return ("%.6g" % x).rstrip("0").rstrip(".")
    return str(x)


def parse_file(path: Path) -> Tuple[str, Dict[str, str], int]:
    """Разобрать JSON и вернуть sample_id, агрегированные значения по препаратам
    и число other_variants.
    """
    with path.open() as f:
        data = json.load(f)

    sample_id: str = data.get("id", "UNKNOWN")

    genes: Dict[str, List[str]] = {drug: [] for drug in DRUGS}
    muts: Dict[str, List[str]] = {drug: [] for drug in DRUGS}
    freqs: Dict[str, List[str]] = {drug: [] for drug in DRUGS}

    for var in data.get("dr_variants", []):
        for drug_info in var.get("drugs", []):
            drug_json = (drug_info.get("drug", "") or "").lower()
            if drug_json not in JSON2COL:
                continue
            conf = (drug_info.get("confidence", "") or "").strip()
            if conf not in {"Assoc w R", "Uncertain significance"}:
                continue
            drug = JSON2COL[drug_json]

            if (gn := var.get("gene_name")):
                genes[drug].append(str(gn))
            if (mut := drug_info.get("original_mutation")):
                if conf == "Uncertain significance":
                    muts[drug].append(f"{mut} [Uncertain significance]")
                else:
                    muts[drug].append(str(mut))
            if (fq := var.get("freq")) is not None and fq != "":
                freqs[drug].append(_to_str(fq))

    # Количество прочих вариантов (не ассоциированных с устойчивостью)
    other_variants_count = len(data.get("other_variants", []))

    row: Dict[str, str] = {}
    for drug in DRUGS:
        row[f"{drug}__gene_name"] = "; ".join(genes[drug])
        row[f"{drug}__Mutation"] = "; ".join(muts[drug])
        row[f"{drug}__Freq"] = "; ".join(freqs[drug])

    return sample_id, row, other_variants_count


def build_dataframe(json_files: List[str]) -> pd.DataFrame:
    rows_out: List[List[str]] = []
    for fp in json_files:
        sample_id, parsed, other_cnt = parse_file(Path(fp))
        row_list: List[str] = [sample_id, str(other_cnt)]
        for drug in DRUGS:
            row_list.extend([
                parsed[f"{drug}__gene_name"],
                parsed[f"{drug}__Mutation"],
                parsed[f"{drug}__Freq"],
            ])
        rows_out.append(row_list)
    df = pd.DataFrame(rows_out, columns=COLUMNS)
    return df


def write_excel(path: Path, df: pd.DataFrame) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "DR"

    # Стиль шапки
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_font = Font(bold=True)

    # 1) Верхняя строка — названия препаратов
    row1 = 1
    col = 1
    ws.cell(row=row1, column=col, value="")  # пусто над Sample
    col += 1
    ws.cell(row=row1, column=col, value="")  # пусто над Other Variants
    col += 1
    for drug in DRUGS:
        ws.merge_cells(start_row=row1, start_column=col, end_row=row1, end_column=col + 2)
        c = ws.cell(row=row1, column=col, value=drug)
        c.alignment = header_align
        c.font = header_font
        col += 3

    # 2) Вторая строка — подзаголовки
    row2 = 2
    ws.cell(row=row2, column=1, value="Sample").alignment = header_align
    ws.cell(row=row2, column=1).font = header_font
    ws.cell(row=row2, column=2, value="Other Variants").alignment = header_align
    ws.cell(row=row2, column=2).font = header_font
    col = 3
    for _ in DRUGS:
        for sub in ("gene_name", "Mutation", "Freq"):
            c = ws.cell(row=row2, column=col, value=sub)
            c.alignment = header_align
            c.font = header_font
            col += 1

    # 3) Данные
    for r, (_, row) in enumerate(df.iterrows(), start=3):
        ws.cell(row=r, column=1, value=row[("", "Sample")])
        ws.cell(row=r, column=2, value=row[("", "Other Variants")])
        col = 3
        for drug in DRUGS:
            g = row[(drug, "gene_name")]
            m = row[(drug, "Mutation")]
            f = row[(drug, "Freq")]
            ws.cell(row=r, column=col, value=g); col += 1
            ws.cell(row=r, column=col, value=m); col += 1
            ws.cell(row=r, column=col, value=f); col += 1

    # Примерная авто-ширина
    for c in range(1, ws.max_column + 1):
        maxlen = 0
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            maxlen = max(maxlen, len(str(v)))
        ws.column_dimensions[get_column_letter(c)].width = min(60, max(10, maxlen + 2))

    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Parse TB‑Profiler JSON и вывести Excel с многоуровневыми колонками: "
            "для каждого препарата gene_name/Mutation/Freq. В таблицу попадают "
            "варианты с confidence = 'Assoc w R' и 'Uncertain significance'. "
            "Частота берётся из dr_variants[].freq. Для 'Uncertain significance' "
            "в колонке Mutation добавляется пометка '[Uncertain significance]'."
        )
    )
    parser.add_argument(
        "-o",
        "--output",
        metavar="XLSX",
        help="Имя выходного Excel-файла (.xlsx). По умолчанию result.xlsx",
    )
    parser.add_argument("json_files", nargs="+", help="Входные файлы *.json")
    args = parser.parse_args()

    df = build_dataframe(args.json_files)

    out_path = Path(args.output or "result.xlsx")
    write_excel(out_path, df)
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()
